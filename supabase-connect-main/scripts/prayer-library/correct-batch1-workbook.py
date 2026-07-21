import argparse
import csv
import hashlib
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_COPY_FIELDS = {
    "Prayer Body": "prayer_body",
    "Import Notes": "import_notes",
    "Source Title": "source_title",
}
PROTECTED_HEADERS = {
    "Prayer Code", "Slug", "Translation Key", "Translation Group ID",
    "Parent Prayer Code", "Prayer Type", "Category Slug", "Sort Order",
}


def text(value):
    return str(value or "").strip()


def normalized(value):
    value = unicodedata.normalize("NFKC", text(value)).casefold()
    return re.sub(r"\s+", " ", value).strip()


def normalized_slug(value):
    return normalized(value).strip("/")


def sheet_rows(sheet):
    values = list(sheet.iter_rows(values_only=True))
    headers = [text(value) for value in values[0]]
    return headers, [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]


def sha256(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest().upper()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--source", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report-json", required=True)
    parser.add_argument("--report-csv", required=True)
    args = parser.parse_args()

    template_book = load_workbook(args.template)
    source_book = load_workbook(args.source, read_only=True, data_only=True)
    template_sheet = template_book["Prayers"]
    source_sheet = source_book["Prayers"]
    template_headers, template_rows = sheet_rows(template_sheet)
    source_headers, source_rows = sheet_rows(source_sheet)
    header_index = {header: index + 1 for index, header in enumerate(template_headers)}

    original_structure = template_book.sheetnames[:]
    original_protected = {
        (row_number, header): template_sheet.cell(row_number, header_index[header]).value
        for row_number in range(2, template_sheet.max_row + 1)
        for header in PROTECTED_HEADERS
    }
    by_code = {normalized(row["Prayer Code"]): (number, row) for number, row in enumerate(template_rows, 2) if text(row["Prayer Code"])}
    by_slug = {normalized_slug(row["Slug"]): (number, row) for number, row in enumerate(template_rows, 2) if text(row["Slug"])}
    by_title_language = {}
    for number, row in enumerate(template_rows, 2):
        by_title_language.setdefault((normalized(row["Title"]), normalized(row["Language"])), []).append((number, row))

    mapping = []
    used_targets = set()
    matched = unmatched = ambiguous = 0
    duplicate_body_titles = {"sala ya mt. inyasi"}

    for source_number, incoming in enumerate(source_rows, 2):
        incoming_code = text(incoming.get("prayer_code"))
        incoming_slug = text(incoming.get("slug"))
        incoming_title = text(incoming.get("title_sw"))
        language = text(incoming.get("language")) or "sw"
        candidates = []
        method = ""

        if normalized(incoming_code) in by_code:
            candidates = [by_code[normalized(incoming_code)]]
            method = "exact existing prayer_code"
        elif normalized_slug(incoming_slug) in by_slug:
            candidates = [by_slug[normalized_slug(incoming_slug)]]
            method = "exact normalized slug"
        else:
            candidates = by_title_language.get((normalized(incoming_title), normalized(language)), [])
            method = "exact normalized Swahili title plus language"

        warning = ""
        action = "skipped"
        confidence = "none"
        target_code = target_title = ""

        if len(candidates) > 1:
            ambiguous += 1
            action = "ambiguous; body left unchanged"
            warning = "Multiple canonical records matched; automatic mapping refused."
        elif not candidates:
            unmatched += 1
            action = "unmatched; body left unchanged"
            warning = "No canonical staging/template record matched the approved exact rules."
            if normalized(incoming_title) in duplicate_body_titles:
                warning += " Body is identical to Sala Baada ya Komunyo and was not copied twice."
        else:
            target_row_number, target = candidates[0]
            target_code = text(target["Prayer Code"])
            target_title = text(target["Title"])
            if target_code in used_targets:
                ambiguous += 1
                action = "duplicate target; body left unchanged"
                warning = "Another incoming row already mapped to this canonical record."
            else:
                used_targets.add(target_code)
                matched += 1
                action = "copied controlled content fields"
                confidence = "high"
                for target_header, source_header in ALLOWED_COPY_FIELDS.items():
                    value = incoming.get(source_header)
                    if text(value):
                        template_sheet.cell(target_row_number, header_index[target_header]).value = value
                template_sheet.cell(target_row_number, header_index["Source Type"]).value = "approved_prayer_book"
                template_sheet.cell(target_row_number, header_index["License Type"]).value = "unknown"
                template_sheet.cell(target_row_number, header_index["Ecclesial Approval Status"]).value = "pending"
                template_sheet.cell(target_row_number, header_index["Status"]).value = "draft"
                template_sheet.cell(target_row_number, header_index["Featured"]).value = False
                if normalized(incoming_title) == "sala baada ya komunyo":
                    warning = "Identical source body also appears as Sala ya Mt. Inyasi; copied only to this canonical record. License remains unknown."
                else:
                    warning = "License remains unknown and blocks publication."

        mapping.append({
            "source_row": source_number,
            "incoming_title": incoming_title,
            "incoming_code": incoming_code,
            "incoming_slug": incoming_slug,
            "matched_prayer_code": target_code,
            "matched_title": target_title,
            "match_method": method if candidates else "none",
            "confidence": confidence,
            "action": action,
            "warning_or_reason": warning,
        })

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    template_book.save(output)

    check_book = load_workbook(output, read_only=False, data_only=False)
    check_sheet = check_book["Prayers"]
    check_headers = [text(cell.value) for cell in check_sheet[1]]
    assert check_book.sheetnames == original_structure, "Worksheet structure changed"
    assert check_headers == template_headers, "Canonical headers changed"
    assert check_sheet.max_row == template_sheet.max_row, "Canonical row count changed"
    for key, value in original_protected.items():
        row_number, header = key
        assert check_sheet.cell(row_number, header_index[header]).value == value, f"Protected field changed: {header} row {row_number}"
    assert all(text(check_sheet.cell(row, header_index["Status"]).value).lower() == "draft" for row in range(2, check_sheet.max_row + 1))
    assert all(check_sheet.cell(row, header_index["Featured"]).value in (False, "false", "False", 0) for row in range(2, check_sheet.max_row + 1))

    report = {
        "template": args.template,
        "source": args.source,
        "output": args.output,
        "output_sha256": sha256(output),
        "source_rows": len(source_rows),
        "safely_matched_rows": matched,
        "unmatched_rows": unmatched,
        "ambiguous_rows": ambiguous,
        "canonical_rows": len(template_rows),
        "worksheet_structure_preserved": True,
        "canonical_headers_preserved": True,
        "protected_fields_preserved": True,
        "license_warning": "Every mapped row retains license_type=unknown; publication remains blocked.",
        "mapping": mapping,
    }
    Path(args.report_json).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    with Path(args.report_csv).open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(mapping[0]))
        writer.writeheader()
        writer.writerows(mapping)
    print(json.dumps({key: value for key, value in report.items() if key != "mapping"}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
