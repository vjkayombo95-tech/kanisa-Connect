import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


READ_ONLY = {"Prayer Code", "Slug", "Translation Key", "Translation Group ID", "Parent Prayer Code", "Prayer Type", "Category Slug", "Sort Order"}
REQUIRED = {"Prayer Code", "Slug", "Title", "Category", "Language", "Status"}
WIDTHS = {
    "Prayer Code": 29, "Slug": 31, "Title": 34, "Parent Prayer Code": 29, "Parent Title": 32,
    "Prayer Type": 21, "Category": 28, "Category Slug": 28, "Sort Order": 12, "Language": 12,
    "Translation Key": 29, "Translation Group ID": 38,
    "Summary": 48, "Prayer Body": 85, "Status": 15, "Visibility": 15, "Featured": 12,
    "Recommended Time": 22, "Scripture Reference": 27, "Liturgical Season": 22, "Audio URL": 38,
    "Author": 28, "Source": 42, "Content Edition": 28, "Content Version": 20,
    "Source Type": 26, "Source Title": 38, "Source Organization": 32, "Source Reference": 38,
    "Source URL": 42, "Source Notes": 48, "Copyright Holder": 32, "Copyright Notice": 48,
    "License Type": 26, "License Reference": 38, "Reviewed By": 28, "Review Date": 16,
    "Ecclesial Approval Status": 28, "Ecclesial Approval Authority": 34,
    "Ecclesial Approval Reference": 38, "Import Notes": 48,
}

INSTRUCTIONS = [
    ("Madhumuni", "Kitabu hiki ni cha kuingiza maandishi ya sala za Kiswahili yaliyopitiwa na yanayoweza kufuatiliwa. Anza kwa staging pekee."),
    ("Vitambulisho", "Prayer Code na Slug havipaswi kubadilishwa. Safu za kijivu zenye alama 'USIBADILISHE' ni vitambulisho vya mfumo."),
    ("Chanzo cha Sala", "Andika jina la kitabu, Misale, Katekisimu, mwongozo wa jimbo, TEC, parokia, au chanzo kingine kinachoweza kuthibitishwa."),
    ("Aina ya Chanzo", "Chagua aina inayofaa kutoka kwenye orodha ya Source Type."),
    ("Haki Miliki na Leseni", "Usidhani kwamba sala yote inayopatikana mtandaoni ni public domain. Ecclesial approval si ruhusa ya hakimiliki. Kwa copyright_restricted, toa ushahidi wa ruhusa au leseni kabla ya uchapishaji."),
    ("Content Edition", "Mfano: Roman Missal, Third Edition; Toleo la 2026; au Kitabu cha Sala, chapa ya pili."),
    ("Content Version", "Mfano: 1.0, 1.1, au 2026.1. Hii si database audit version ya content_versions."),
    ("Translation Key", "Tafsiri zote za sala moja hushiriki Translation Key moja. Usibadilishe kwa mkono kwenye workbook iliyotengenezwa."),
    ("Mapitio ya Tafsiri", "Kila tafsiri lazima ipitiwe kivyake. Idhini ya Kiswahili haiidhinishi Kiingereza au Kilatini; chanzo na hakimiliki vinaweza kutofautiana kwa lugha."),
    ("Hali ya kwanza", "Weka Status = draft wakati wa kuingiza maandishi kwa mara ya kwanza."),
    ("Mapitio", "Tumia review baada ya uhakiki wa ndani wa tahajia, maana, chanzo na mpangilio."),
    ("Uchapishaji", "Tumia published tu baada ya idhini ya maudhui/ki-kanisa na browser UAT. Uingizaji wa kwanza wa CLI utaendelea kuweka draft isipokuwa ruhusa ya kuchapisha imetolewa wazi."),
    ("Maelezo ya chanzo", "Source itaje kitabu cha sala, Misale, Katekisimu, maandishi ya jimbo, au rejea nyingine yenye mamlaka."),
    ("Tarehe", "Review Date lazima iwe katika muundo YYYY-MM-DD."),
    ("Mpangilio", "Parent Prayer Code na Sort Order havipaswi kubadilishwa bila mapitio ya muundo wa mkusanyo."),
    ("Sehemu tupu", "Prayer Body ikiwa tupu haitaondoa maandishi yaliyopo; kwa rekodi mpya za seed, sala itaendelea kutopatikana kwa waumini."),
    ("Uhifadhi", "Seli tupu hazifuti maudhui yaliyopo. Jaza seli tu unapokusudia kubadilisha thamani."),
    ("Mtiririko", "Kamilisha workbook, endesha dry-run, rekebisha makosa yote, kisha ingiza workbook hiyo kwenye staging kabla ya uzalishaji."),
    ("Mafungu ya Rozari", "Mkusanyo mkuu wenye watoto watano waliopangwa. Usibadilishe Parent Prayer Code au Sort Order."),
    ("Njia ya Msalaba", "Mkusanyo mkuu wenye watoto kumi na sita waliopangwa."),
    ("Sala za Misa Takatifu", "Mkusanyo mkuu wenye watoto wanane waliopangwa."),
]


def style_header(cell, read_only=False, required=False):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if read_only:
        cell.fill = PatternFill("solid", fgColor="59636E")
        cell.comment = Comment("USIBADILISHE — kitambulisho hiki kinalindwa na mfumo.", "Kanisa Connect")
    elif required:
        cell.fill = PatternFill("solid", fgColor="B7791F")
    else:
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def add_table(ws, name, rows, columns):
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(2, ws.max_row)}"
    if rows:
        table = Table(displayName=name, ref=f"A1:{get_column_letter(len(columns))}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)
    for index, column in enumerate(columns, 1):
        style_header(ws.cell(1, index), column in READ_ONLY, column in REQUIRED)
        ws.column_dimensions[get_column_letter(index)].width = WIDTHS.get(column, max(16, len(column) + 3))
    ws.row_dimensions[1].height = 34


def build_prayers(workbook, data):
    ws = workbook.active
    ws.title = "Prayers"
    headers = data["headers"]
    add_table(ws, "PrayerContentRows", data["prayers"], headers)
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.password = "review"
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = False
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 54
        for column, header in enumerate(headers, 1):
            cell = ws.cell(row, column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=header in READ_ONLY)
            if header in READ_ONLY:
                cell.fill = PatternFill("solid", fgColor="E2E8F0")
            elif header in REQUIRED:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")

    list_columns = {
        "Status": "'Validation Lists'!$A$2:$A$5",
        "Visibility": "'Validation Lists'!$B$2:$B$4",
        "Featured": "'Validation Lists'!$C$2:$C$3",
        "Ecclesial Approval Status": "'Validation Lists'!$D$2:$D$6",
        "Language": "'Validation Lists'!$E$2:$E$4",
        "Source Type": "'Validation Lists'!$F$2:$F$12",
        "License Type": "'Validation Lists'!$G$2:$G$8",
    }
    for header, formula in list_columns.items():
        column = headers.index(header) + 1
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = f"Chagua thamani halali ya {header}."
        validation.errorTitle = "Thamani si halali"
        validation.prompt = f"Chagua {header} kutoka orodha."
        validation.promptTitle = "Kanisa Connect"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{max(ws.max_row, 500)}")

    review_column = headers.index("Review Date") + 1
    review_validation = DataValidation(type="custom", formula1=f'=OR({get_column_letter(review_column)}2="",AND(LEN({get_column_letter(review_column)}2)=10,MID({get_column_letter(review_column)}2,5,1)="-",MID({get_column_letter(review_column)}2,8,1)="-"))', allow_blank=True)
    review_validation.error = "Tumia YYYY-MM-DD."
    review_validation.showErrorMessage = True
    ws.add_data_validation(review_validation)
    review_validation.add(f"{get_column_letter(review_column)}2:{get_column_letter(review_column)}{max(ws.max_row, 500)}")
    body_column = get_column_letter(headers.index("Prayer Body") + 1)
    ws.conditional_formatting.add(f"{body_column}2:{body_column}{ws.max_row}", FormulaRule(formula=[f'LEN(TRIM({body_column}2))=0'], fill=PatternFill("solid", fgColor="FCE8E6")))


def build_instructions(workbook):
    ws = workbook.create_sheet("Instructions")
    ws.append(["Mada", "Maelekezo ya kuingiza maudhui ya sala"])
    for title, text in INSTRUCTIONS:
        ws.append([title, text])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 115
    for cell in ws[1]:
        style_header(cell)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True, color="1F4E78")
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 42
    ws.sheet_view.showGridLines = False


def build_categories(workbook, data):
    ws = workbook.create_sheet("Categories")
    rows = [{"Category": item["name"], "Category Slug": item["slug"]} for item in data["categories"]]
    add_table(ws, "PrayerCategories", rows, ["Category", "Category Slug"])


def build_collections(workbook, data):
    ws = workbook.create_sheet("Collections")
    columns = ["Parent Prayer Code", "Parent Title", "Child Prayer Code", "Child Title", "Child Type", "Sort Order"]
    rows = [{
        "Parent Prayer Code": item["parentPrayerCode"], "Parent Title": item["parentTitle"],
        "Child Prayer Code": item["childPrayerCode"], "Child Title": item["childTitle"],
        "Child Type": item["childType"], "Sort Order": item["sortOrder"],
    } for item in data["collections"]]
    add_table(ws, "PrayerCollections", rows, columns)
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 31
    ws.column_dimensions["D"].width = 36


def build_validation_lists(workbook):
    ws = workbook.create_sheet("Validation Lists")
    values = {
        "Status": ["draft", "review", "published", "archived"],
        "Visibility": ["member", "admin", "private"],
        "Featured": ["true", "false"],
        "Ecclesial Approval Status": ["pending", "under_review", "approved", "rejected", "revision_required"],
        "Language": ["sw", "en", "la"],
        "Source Type": ["roman_missal", "catechism", "bishops_conference", "diocesan_publication", "parish_publication", "approved_prayer_book", "scripture", "public_domain", "original_parish_content", "user_submitted", "other"],
        "License Type": ["public_domain", "permission_granted", "licensed", "attribution_required", "internal_church_use", "copyright_restricted", "unknown"],
    }
    ws.append(list(values.keys()))
    for index in range(max(len(items) for items in values.values())):
        ws.append([items[index] if index < len(items) else "" for items in values.values()])
    for cell in ws[1]:
        style_header(cell)
    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 30
    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    workbook = Workbook()
    build_prayers(workbook, data)
    build_instructions(workbook)
    build_categories(workbook, data)
    build_collections(workbook, data)
    build_validation_lists(workbook)
    workbook.calculation.fullCalcOnLoad = True
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Formatted workbook written: {output}")


if __name__ == "__main__":
    main()
