from __future__ import annotations

import csv
import hashlib
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from .corpus import BENCHMARK_CORPUS, BenchmarkChapter
from .models import Transcript, VerseBoundary, WordTiming


HEADER_ALIASES = {
    "book": {"book", "kitabu"},
    "chapter": {"chapter", "sura"},
    "chapter_id": {"chapter_id", "osis_chapter", "id"},
    "verse": {"verse", "mstari"},
    "verse_text": {"verse_text", "text", "swahili", "swahili_text", "reference_text", "golden_text"},
    "word": {"word", "neno"},
    "start_ms": {"start_ms", "word_start_ms", "start", "start_time_ms"},
    "end_ms": {"end_ms", "word_end_ms", "end", "end_time_ms"},
    "confidence": {"confidence", "score", "word_confidence"},
    "verse_start_ms": {"verse_start_ms", "boundary_start_ms"},
    "verse_end_ms": {"verse_end_ms", "boundary_end_ms"},
    "verse_confidence": {"verse_confidence", "boundary_confidence"},
}


class GoldenReferenceImportError(ValueError):
    """Raised when a golden reference workbook cannot be normalized."""


class GoldenReferenceSpreadsheetImporter:
    """Parse manually corrected Swahili golden references without running ASR."""

    def import_file(self, path: str | Path) -> list[Transcript]:
        source = Path(path)
        if source.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = list(self._read_xlsx(source))
        elif source.suffix.lower() == ".csv":
            rows = list(self._read_csv(source))
        else:
            raise GoldenReferenceImportError(f"Unsupported golden reference file type: {source.suffix}")
        return self.from_rows(rows, source_name=source.name, source_hash=self.sha256(source))

    def from_rows(self, rows: Iterable[dict[str, Any]], *, source_name: str | None = None, source_hash: str | None = None) -> list[Transcript]:
        normalized_rows = [
            {**self._normalize_row(row), "_row_order": index}
            for index, row in enumerate(rows)
            if any(str(value).strip() for value in row.values())
        ]
        if not normalized_rows:
            raise GoldenReferenceImportError("Golden reference file contains no data rows.")

        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in normalized_rows:
            chapter_id = str(row.get("chapter_id") or self._chapter_id(row)).strip()
            row["chapter_id"] = chapter_id
            grouped[chapter_id].append(row)

        transcripts = [self._transcript_from_rows(chapter_id, chapter_rows, source_name, source_hash) for chapter_id, chapter_rows in grouped.items()]
        return sorted(transcripts, key=lambda item: item.chapter_id)

    def _read_csv(self, path: Path) -> Iterable[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            yield from csv.DictReader(handle)

    def _read_xlsx(self, path: Path) -> Iterable[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - depends on local optional package
            raise GoldenReferenceImportError("Install openpyxl to import .xlsx golden references.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            except StopIteration:
                continue
            for values in rows:
                yield {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}

    def _normalize_row(self, row: dict[str, Any]) -> dict[str, Any]:
        by_normalized_key = {_key(key): value for key, value in row.items()}
        normalized: dict[str, Any] = {}
        for canonical, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in by_normalized_key:
                    normalized[canonical] = by_normalized_key[alias]
                    break
        return normalized

    def _chapter_id(self, row: dict[str, Any]) -> str:
        book = str(row.get("book", "")).strip()
        chapter = _int(row.get("chapter"))
        for benchmark in BENCHMARK_CORPUS:
            if benchmark.book.casefold() == book.casefold() and benchmark.chapter == chapter:
                return benchmark.id
        raise GoldenReferenceImportError(f"Cannot resolve benchmark chapter for row: book={book!r}, chapter={chapter!r}")

    def _transcript_from_rows(
        self,
        chapter_id: str,
        rows: list[dict[str, Any]],
        source_name: str | None,
        source_hash: str | None,
    ) -> Transcript:
        benchmark = _benchmark_by_id(chapter_id)
        verse_rows = sorted(rows, key=lambda row: (_int(row.get("verse")), _int(row.get("_row_order"))))
        verse_texts: dict[int, str] = {}
        words: list[WordTiming] = []
        boundaries: dict[int, VerseBoundary] = {}

        for row in verse_rows:
            verse = _int(row.get("verse"))
            text = _text(row.get("verse_text"))
            if verse and text and verse not in verse_texts:
                verse_texts[verse] = text
            word = _text(row.get("word"))
            if word:
                words.append(
                    WordTiming(
                        word=word,
                        start_ms=_optional_int(row.get("start_ms")),
                        end_ms=_optional_int(row.get("end_ms")),
                        confidence=_optional_float(row.get("confidence")),
                        verse=verse or None,
                    )
                )
            if verse and row.get("verse_start_ms") not in (None, "") and row.get("verse_end_ms") not in (None, ""):
                boundaries[verse] = VerseBoundary(
                    verse=verse,
                    start_ms=_int(row.get("verse_start_ms")),
                    end_ms=_int(row.get("verse_end_ms")),
                    confidence=_optional_float(row.get("verse_confidence")),
                )

        if not verse_texts and not words:
            raise GoldenReferenceImportError(f"{chapter_id} does not contain verse text or word rows.")

        return Transcript(
            chapter_id=chapter_id,
            text=" ".join(verse_texts[verse] for verse in sorted(verse_texts)),
            words=words,
            verse_boundaries=[boundaries[verse] for verse in sorted(boundaries)],
            metadata={
                "book": benchmark.book,
                "chapter": benchmark.chapter,
                "translation_code": "sw-biblica",
                "source_name": source_name,
                "source_hash": source_hash,
                "status": "golden_reference",
            },
        )

    @staticmethod
    def sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()


def _benchmark_by_id(chapter_id: str) -> BenchmarkChapter:
    for benchmark in BENCHMARK_CORPUS:
        if benchmark.id == chapter_id:
            return benchmark
    raise GoldenReferenceImportError(f"Golden reference chapter is not in the fixed benchmark corpus: {chapter_id}")


def _key(value: str) -> str:
    return value.strip().casefold().replace(" ", "_").replace("-", "_")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _int(value: Any) -> int:
    if value in (None, ""):
        return 0
    return int(float(str(value).strip()))


def _optional_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return _int(value)


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(str(value).strip())
